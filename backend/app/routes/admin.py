from __future__ import annotations

import hashlib
import hmac
import csv
import io
import json
import time
from base64 import urlsafe_b64decode, urlsafe_b64encode
from datetime import date, datetime, timedelta, timezone
from typing import Literal

import httpx
from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import Date as SqlDate, cast, func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.config import get_settings, update_settings
from app.db.session import get_db
from app.models import (
    LearnerDailyUsage,
    LearnerProgress,
    LearnerUser,
    PaymentRequest,
    PaymentStatus,
    PointsEvent,
    WordItem,
)
from app.routes.words import word_payload
from app.services.payments import approve_payment, cancel_payment

router = APIRouter(prefix="/api/admin", tags=["admin"])
_ADMIN_SESSION_TTL = 7 * 24 * 3600


def _b64encode(data: bytes) -> str:
    return urlsafe_b64encode(data).decode().rstrip("=")


def _b64decode(data: str) -> bytes:
    padding = "=" * (-len(data) % 4)
    return urlsafe_b64decode(data + padding)


def _admin_login_secret() -> str:
    settings = get_settings()
    return settings.admin_password or settings.admin_approval_token


def _create_admin_session() -> str:
    settings = get_settings()
    now = int(time.time())
    payload = {"sub": "admin", "iat": now, "exp": now + _ADMIN_SESSION_TTL}
    encoded_payload = _b64encode(json.dumps(payload, separators=(",", ":")).encode())
    signature = hmac.new(settings.secret_key.encode(), encoded_payload.encode(), hashlib.sha256).digest()
    return f"{encoded_payload}.{_b64encode(signature)}"


def _verify_admin_session(token: str) -> None:
    try:
        encoded_payload, received_signature = token.split(".", 1)
    except ValueError as exc:
        raise HTTPException(status_code=403, detail="Invalid admin session") from exc

    settings = get_settings()
    expected_signature = hmac.new(settings.secret_key.encode(), encoded_payload.encode(), hashlib.sha256).digest()
    if not hmac.compare_digest(_b64encode(expected_signature), received_signature):
        raise HTTPException(status_code=403, detail="Invalid admin session")

    try:
        payload = json.loads(_b64decode(encoded_payload))
        if payload.get("sub") != "admin" or int(time.time()) > payload.get("exp", 0):
            raise HTTPException(status_code=403, detail="Admin session expired")
    except (KeyError, ValueError, TypeError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=403, detail="Invalid admin session") from exc


class AdminLoginRequest(BaseModel):
    password: str = Field(min_length=1)


class AdminWordPayload(BaseModel):
    word: str = Field(min_length=1, max_length=255)
    word_type: str = Field(min_length=1, max_length=80)
    phonetic: str = Field(min_length=1, max_length=255)
    english_definition: str = Field(min_length=1)
    uzbek_definition: str = Field(min_length=1)
    english_example: str = Field(min_length=1)
    uzbek_example: str = Field(min_length=1)
    level: str = Field(min_length=1, max_length=32)
    topic: str = Field(default="General", min_length=1, max_length=120)
    collection: str = Field(default="Daily Vocabulary", min_length=1, max_length=160)
    tags: str = ""
    collocations: str = ""
    common_mistake: str = ""
    writing_prompt: str = ""
    difficulty_order: int = Field(default=0, ge=0)
    audio_url: str | None = Field(default=None, max_length=500)
    is_active: bool = True


class AdminWordPatch(BaseModel):
    word: str | None = Field(default=None, min_length=1, max_length=255)
    word_type: str | None = Field(default=None, min_length=1, max_length=80)
    phonetic: str | None = Field(default=None, min_length=1, max_length=255)
    english_definition: str | None = Field(default=None, min_length=1)
    uzbek_definition: str | None = Field(default=None, min_length=1)
    english_example: str | None = Field(default=None, min_length=1)
    uzbek_example: str | None = Field(default=None, min_length=1)
    level: str | None = Field(default=None, min_length=1, max_length=32)
    topic: str | None = Field(default=None, min_length=1, max_length=120)
    collection: str | None = Field(default=None, min_length=1, max_length=160)
    tags: str | None = None
    collocations: str | None = None
    common_mistake: str | None = None
    writing_prompt: str | None = None
    difficulty_order: int | None = Field(default=None, ge=0)
    audio_url: str | None = Field(default=None, max_length=500)
    is_active: bool | None = None


class AdminImportUrlRequest(BaseModel):
    url: str = Field(min_length=1)
    default_active: bool = True


def require_admin_token(
    authorization: str | None = Header(default=None, alias="Authorization"),
    x_admin_token: str | None = Header(default=None, alias="X-Admin-Token"),
) -> None:
    settings = get_settings()
    if authorization and authorization.lower().startswith("bearer "):
        _verify_admin_session(authorization.split(" ", 1)[1].strip())
        return
    if settings.admin_approval_token and x_admin_token == settings.admin_approval_token:
        return
    raise HTTPException(status_code=403, detail="Invalid admin session")


def admin_word_payload(word: WordItem) -> dict:
    payload = word_payload(word)
    payload["is_active"] = word.is_active
    payload["created_at"] = word.created_at.isoformat() if word.created_at else None
    return payload


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _bool_cell(value: object, default: bool = True) -> bool:
    text_value = _clean_cell(value).lower()
    if not text_value:
        return default
    return text_value in {"1", "true", "yes", "y", "published", "active"}


def _int_cell(value: object, default: int = 0) -> int:
    try:
        return max(0, int(float(_clean_cell(value))))
    except ValueError:
        return default


def _word_from_row(row: dict[str, object], default_active: bool) -> tuple[WordItem | None, str | None]:
    normalized = {key.strip().lower(): value for key, value in row.items() if key}
    word = _clean_cell(normalized.get("word"))
    if not word:
        return None, "Missing word"

    english_definition = _clean_cell(normalized.get("english_definition") or normalized.get("definition"))
    uzbek_definition = _clean_cell(normalized.get("uzbek_definition") or normalized.get("uzbek_meaning") or normalized.get("meaning"))
    english_example = _clean_cell(normalized.get("english_example") or normalized.get("example"))
    uzbek_example = _clean_cell(normalized.get("uzbek_example") or normalized.get("translation"))
    missing = [
        name
        for name, value in {
            "english_definition": english_definition,
            "uzbek_definition": uzbek_definition,
            "english_example": english_example,
            "uzbek_example": uzbek_example,
        }.items()
        if not value
    ]
    if missing:
        return None, f"{word}: missing {', '.join(missing)}"

    return (
        WordItem(
            word=word,
            word_type=_clean_cell(normalized.get("word_type") or normalized.get("type")) or "word",
            phonetic=_clean_cell(normalized.get("phonetic")) or "-",
            english_definition=english_definition,
            uzbek_definition=uzbek_definition,
            english_example=english_example,
            uzbek_example=uzbek_example,
            level=_clean_cell(normalized.get("level")) or "A1",
            topic=_clean_cell(normalized.get("topic")) or "General",
            collection=_clean_cell(normalized.get("collection")) or "Daily Vocabulary",
            tags=_clean_cell(normalized.get("tags")),
            collocations=_clean_cell(normalized.get("collocations")),
            common_mistake=_clean_cell(normalized.get("common_mistake")),
            writing_prompt=_clean_cell(normalized.get("writing_prompt")),
            difficulty_order=_int_cell(normalized.get("difficulty_order")),
            audio_url=_clean_cell(normalized.get("audio_url")) or None,
            is_active=_bool_cell(normalized.get("is_active"), default_active),
        ),
        None,
    )


def _rows_from_csv(content: bytes) -> list[dict[str, object]]:
    text_content = content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text_content)))


def _rows_from_xlsx(content: bytes) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean_cell(cell).lower() for cell in rows[0]]
    parsed_rows = []
    for values in rows[1:]:
        parsed_rows.append({headers[index]: value for index, value in enumerate(values) if index < len(headers)})
    return parsed_rows


def _import_rows(rows: list[dict[str, object]], default_active: bool, db: Session) -> dict:
    imported = 0
    skipped: list[str] = []
    for index, row in enumerate(rows, start=2):
        word, error = _word_from_row(row, default_active)
        if error or not word:
            skipped.append(f"Row {index}: {error or 'Invalid row'}")
            continue
        db.add(word)
        imported += 1
    db.commit()
    return {"imported": imported, "skipped": skipped[:25], "skipped_count": len(skipped)}


@router.post("/login")
def admin_login(payload: AdminLoginRequest) -> dict:
    expected_password = _admin_login_secret()
    if not expected_password:
        raise HTTPException(status_code=403, detail="Admin password is not configured")
    if not hmac.compare_digest(payload.password, expected_password):
        raise HTTPException(status_code=403, detail="Invalid admin password")
    return {"token": _create_admin_session()}


@router.get("/summary")
def admin_summary(db: Session = Depends(get_db), _: None = Depends(require_admin_token)) -> dict:
    total_words = db.scalar(select(func.count(WordItem.id))) or 0
    published_words = db.scalar(select(func.count(WordItem.id)).where(WordItem.is_active.is_(True))) or 0
    draft_words = total_words - published_words
    total_users = db.scalar(select(func.count(LearnerUser.id))) or 0
    premium_users = db.scalar(select(func.count(LearnerUser.id)).where(LearnerUser.tier == "paid")) or 0
    pending_payments = (
        db.scalar(
            select(func.count(PaymentRequest.id)).where(
                PaymentRequest.status.in_([PaymentStatus.PENDING.value, PaymentStatus.SUBMITTED.value])
            )
        )
        or 0
    )
    recent_words = list(db.scalars(select(WordItem).order_by(WordItem.created_at.desc()).limit(5)))
    return {
        "stats": {
            "total_words": total_words,
            "published_words": published_words,
            "draft_words": draft_words,
            "total_users": total_users,
            "premium_users": premium_users,
            "pending_payments": pending_payments,
        },
        "recent_words": [admin_word_payload(word) for word in recent_words],
    }


@router.get("/words")
def admin_words(
    search: str = "",
    status: Literal["all", "published", "draft"] = "all",
    level: str = "",
    word_type: str = "",
    limit: int = Query(default=100, ge=1, le=300),
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    query = select(WordItem)
    if search.strip():
        term = f"%{search.strip()}%"
        query = query.where(
            or_(
                WordItem.word.ilike(term),
                WordItem.uzbek_definition.ilike(term),
                WordItem.english_definition.ilike(term),
                WordItem.topic.ilike(term),
                WordItem.collection.ilike(term),
                WordItem.tags.ilike(term),
            )
        )
    if status == "published":
        query = query.where(WordItem.is_active.is_(True))
    elif status == "draft":
        query = query.where(WordItem.is_active.is_(False))
    if level.strip():
        query = query.where(WordItem.level == level.strip())
    if word_type.strip():
        query = query.where(WordItem.word_type == word_type.strip())

    rows = list(
        db.scalars(
            query.order_by(
                WordItem.difficulty_order.asc(),
                WordItem.created_at.desc(),
                WordItem.id.desc(),
            ).limit(limit)
        )
    )
    return {"items": [admin_word_payload(word) for word in rows]}


@router.post("/words")
def create_admin_word(
    payload: AdminWordPayload,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    word = WordItem(**payload.model_dump())
    db.add(word)
    db.commit()
    db.refresh(word)
    return admin_word_payload(word)


@router.post("/words/import-file")
async def import_admin_words_file(
    upload: UploadFile = File(...),
    default_active: bool = Form(default=True),
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    filename = (upload.filename or "").lower()
    content = await upload.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if filename.endswith(".csv"):
        rows = _rows_from_csv(content)
    elif filename.endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(content)
    else:
        raise HTTPException(status_code=400, detail="Upload CSV or XLSX file")
    return _import_rows(rows, default_active, db)


@router.post("/words/import-url")
async def import_admin_words_url(
    payload: AdminImportUrlRequest,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            response = await client.get(payload.url)
            response.raise_for_status()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=400, detail="Could not read Google Sheets CSV link") from exc
    return _import_rows(_rows_from_csv(response.content), payload.default_active, db)


@router.patch("/words/{word_id}")
def update_admin_word(
    word_id: int,
    payload: AdminWordPatch,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    word = db.get(WordItem, word_id)
    if not word:
        raise HTTPException(status_code=404, detail="Word not found")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(word, key, value)
    db.commit()
    db.refresh(word)
    return admin_word_payload(word)


@router.delete("/words/{word_id}")
def disable_admin_word(
    word_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    word = db.get(WordItem, word_id)
    if not word:
        raise HTTPException(status_code=404, detail="Word not found")
    word.is_active = False
    db.commit()
    db.refresh(word)
    return admin_word_payload(word)


# ───────────── Pydantic models for new endpoints ─────────────

class AdminUserPatch(BaseModel):
    tier: Literal["free", "paid"] | None = None
    premium_until: str | None = None  # ISO datetime string, or empty string to clear


class AdminSettingsPatch(BaseModel):
    free_daily_word_limit: int | None = Field(default=None, ge=1, le=100)
    manual_payment_amount_uzs: int | None = Field(default=None, ge=1000)
    manual_payment_plan_days: int | None = Field(default=None, ge=1, le=365)
    manual_payment_card_label: str | None = Field(default=None, min_length=1, max_length=200)


# ───────────── Helpers ─────────────

def _user_payload(user: LearnerUser, learned_count: int = 0, total_points: int = 0) -> dict:
    return {
        "id": user.id,
        "display_name": user.display_name,
        "username": user.username,
        "tier": user.tier,
        "premium_until": user.premium_until.isoformat() if user.premium_until else None,
        "streak_days": user.streak_days,
        "last_seen_at": user.last_seen_at.isoformat() if user.last_seen_at else None,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "learned_count": learned_count,
        "total_points": total_points,
    }


def _payment_admin_payload(payment: PaymentRequest) -> dict:
    u = payment.user
    return {
        "id": payment.id,
        "code": payment.code,
        "user_id": payment.user_id,
        "user_display_name": u.display_name if u else "—",
        "user_username": u.username if u else None,
        "plan": payment.plan,
        "plan_days": payment.plan_days,
        "amount_uzs": payment.amount_uzs,
        "status": payment.status,
        "admin_note": payment.admin_note,
        "created_at": payment.created_at.isoformat() if payment.created_at else None,
        "submitted_at": payment.submitted_at.isoformat() if payment.submitted_at else None,
        "approved_at": payment.approved_at.isoformat() if payment.approved_at else None,
        "cancelled_at": payment.cancelled_at.isoformat() if payment.cancelled_at else None,
        "expires_at": payment.expires_at.isoformat() if payment.expires_at else None,
    }


def _fill_date_series(rows: list[tuple[date, int]], days: int) -> list[dict]:
    data_map = {str(r[0]): r[1] for r in rows}
    today = datetime.now(timezone.utc).date()
    return [
        {"date": str(today - timedelta(days=days - 1 - i)), "count": data_map.get(str(today - timedelta(days=days - 1 - i)), 0)}
        for i in range(days)
    ]


# ───────────── Users endpoints ─────────────

@router.get("/users")
def admin_users_list(
    search: str = "",
    tier: str = "",
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    learned_subq = (
        select(LearnerProgress.user_id, func.count(LearnerProgress.id).label("learned_count"))
        .where(LearnerProgress.status.in_(["learned", "mastered"]))
        .group_by(LearnerProgress.user_id)
        .subquery()
    )
    points_subq = (
        select(PointsEvent.user_id, func.coalesce(func.sum(PointsEvent.points), 0).label("total_points"))
        .group_by(PointsEvent.user_id)
        .subquery()
    )
    base_q = (
        select(
            LearnerUser,
            func.coalesce(learned_subq.c.learned_count, 0).label("learned_count"),
            func.coalesce(points_subq.c.total_points, 0).label("total_points"),
        )
        .outerjoin(learned_subq, LearnerUser.id == learned_subq.c.user_id)
        .outerjoin(points_subq, LearnerUser.id == points_subq.c.user_id)
    )
    if search.strip():
        term = f"%{search.strip()}%"
        base_q = base_q.where(or_(LearnerUser.display_name.ilike(term), LearnerUser.username.ilike(term)))
    if tier in ("free", "paid"):
        base_q = base_q.where(LearnerUser.tier == tier)

    total = db.scalar(select(func.count(LearnerUser.id))) or 0
    rows = db.execute(base_q.order_by(LearnerUser.created_at.desc()).limit(limit).offset(offset)).all()
    return {
        "total": total,
        "items": [_user_payload(row[0], int(row[1]), int(row[2])) for row in rows],
    }


@router.get("/users/{user_id}")
def admin_user_detail(
    user_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    user = db.get(LearnerUser, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    learned_count = db.scalar(
        select(func.count(LearnerProgress.id)).where(
            LearnerProgress.user_id == user_id,
            LearnerProgress.status.in_(["learned", "mastered"]),
        )
    ) or 0
    mastered_count = db.scalar(
        select(func.count(LearnerProgress.id)).where(
            LearnerProgress.user_id == user_id,
            LearnerProgress.status == "mastered",
        )
    ) or 0
    total_points = db.scalar(
        select(func.coalesce(func.sum(PointsEvent.points), 0)).where(PointsEvent.user_id == user_id)
    ) or 0
    quiz_count = db.scalar(
        select(func.count(LearnerProgress.id)).where(
            LearnerProgress.user_id == user_id,
            LearnerProgress.times_answered > 0,
        )
    ) or 0
    correct_answers = db.scalar(
        select(func.coalesce(func.sum(LearnerProgress.times_correct), 0)).where(LearnerProgress.user_id == user_id)
    ) or 0
    total_answers = db.scalar(
        select(func.coalesce(func.sum(LearnerProgress.times_answered), 0)).where(LearnerProgress.user_id == user_id)
    ) or 0
    accuracy = round(correct_answers / total_answers * 100) if total_answers else 0
    base = _user_payload(user, int(learned_count), int(total_points))
    base.update({
        "mastered_count": int(mastered_count),
        "quiz_attempted": int(quiz_count),
        "quiz_accuracy": accuracy,
    })
    return base


@router.patch("/users/{user_id}")
def admin_update_user(
    user_id: int,
    payload: AdminUserPatch,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    user = db.get(LearnerUser, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.tier is not None:
        user.tier = payload.tier
    if payload.premium_until is not None:
        if payload.premium_until == "":
            user.premium_until = None
        else:
            try:
                user.premium_until = datetime.fromisoformat(payload.premium_until)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail="Invalid datetime format") from exc
    db.commit()
    db.refresh(user)
    return _user_payload(user)


# ───────────── Payments endpoints ─────────────

@router.get("/payments")
def admin_payments_list(
    status: str = "all",
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    q = select(PaymentRequest).options(selectinload(PaymentRequest.user))
    if status != "all":
        q = q.where(PaymentRequest.status == status)

    total = db.scalar(select(func.count(PaymentRequest.id)).where(
        PaymentRequest.status == status if status != "all" else True
    )) or 0
    payments = list(db.scalars(q.order_by(PaymentRequest.created_at.desc()).limit(limit).offset(offset)))

    pending_count = db.scalar(select(func.count(PaymentRequest.id)).where(PaymentRequest.status == PaymentStatus.PENDING.value)) or 0
    submitted_count = db.scalar(select(func.count(PaymentRequest.id)).where(PaymentRequest.status == PaymentStatus.SUBMITTED.value)) or 0
    approved_count = db.scalar(select(func.count(PaymentRequest.id)).where(PaymentRequest.status == PaymentStatus.APPROVED.value)) or 0
    cancelled_count = db.scalar(select(func.count(PaymentRequest.id)).where(PaymentRequest.status == PaymentStatus.CANCELLED.value)) or 0

    return {
        "total": total,
        "counts": {
            "pending": int(pending_count),
            "submitted": int(submitted_count),
            "approved": int(approved_count),
            "cancelled": int(cancelled_count),
        },
        "items": [_payment_admin_payload(p) for p in payments],
    }


@router.post("/payments/{code}/approve")
def admin_approve_payment_session(
    code: str,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    payment = approve_payment(db, code, admin_id="admin-panel")
    return _payment_admin_payload(payment)


@router.post("/payments/{code}/cancel")
def admin_cancel_payment_session(
    code: str,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    payment = cancel_payment(db, code, admin_id="admin-panel")
    return _payment_admin_payload(payment)


# ───────────── Analytics endpoints ─────────────

@router.get("/analytics")
def admin_analytics(
    days: int = Query(default=30, ge=7, le=90),
    db: Session = Depends(get_db),
    _: None = Depends(require_admin_token),
) -> dict:
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    cutoff_date = cutoff.date()

    signup_rows = db.execute(
        select(cast(LearnerUser.created_at, SqlDate).label("day"), func.count(LearnerUser.id).label("cnt"))
        .where(LearnerUser.created_at >= cutoff)
        .group_by(cast(LearnerUser.created_at, SqlDate))
        .order_by(cast(LearnerUser.created_at, SqlDate))
    ).all()

    dau_rows = db.execute(
        select(LearnerDailyUsage.usage_date.label("day"), func.count(LearnerDailyUsage.user_id.distinct()).label("cnt"))
        .where(LearnerDailyUsage.usage_date >= cutoff_date)
        .group_by(LearnerDailyUsage.usage_date)
        .order_by(LearnerDailyUsage.usage_date)
    ).all()

    return {
        "signups": _fill_date_series([(r[0], r[1]) for r in signup_rows], days),
        "dau": _fill_date_series([(r[0], r[1]) for r in dau_rows], days),
    }


# ───────────── Settings endpoints ─────────────

@router.get("/settings")
def admin_get_settings(_: None = Depends(require_admin_token)) -> dict:
    s = get_settings()
    return {
        "free_daily_word_limit": s.free_daily_word_limit,
        "manual_payment_amount_uzs": s.manual_payment_amount_uzs,
        "manual_payment_plan_days": s.manual_payment_plan_days,
        "manual_payment_card_label": s.manual_payment_card_label,
    }


@router.patch("/settings")
def admin_patch_settings(
    payload: AdminSettingsPatch,
    _: None = Depends(require_admin_token),
) -> dict:
    overrides = payload.model_dump(exclude_none=True)
    if overrides:
        update_settings(**overrides)
    return admin_get_settings(_=None)
