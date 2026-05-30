from __future__ import annotations

import csv
import io
from typing import Any

from sqlalchemy.orm import Session

from app.models import WordItem, WordQualityStatus

REQUIRED_FIELDS = ("english_definition", "uzbek_definition", "english_example", "uzbek_example")
ALIASES: dict[str, tuple[str, ...]] = {
    "english_definition": ("english_definition", "definition"),
    "uzbek_definition": ("uzbek_definition", "uzbek_meaning", "meaning"),
    "english_example": ("english_example", "example"),
    "uzbek_example": ("uzbek_example", "translation"),
    "word_type": ("word_type", "type"),
}
TRUTHY = {"1", "true", "yes", "y", "published", "active"}
QUALITY_STATUSES = {status.value for status in WordQualityStatus}


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _get(row: dict, key: str) -> str:
    for alias in ALIASES.get(key, (key,)):
        if value := _clean(row.get(alias)):
            return value
    return ""


def _to_bool(value: Any, default: bool = True) -> bool:
    text = _clean(value).lower()
    return default if not text else text in TRUTHY


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return max(0, int(float(_clean(value))))
    except ValueError:
        return default


def _quality_status(row: dict, default_active: bool) -> str:
    explicit = _clean(row.get("quality_status") or row.get("status")).lower()
    if explicit in QUALITY_STATUSES:
        return explicit
    return WordQualityStatus.PUBLISHED.value if default_active else WordQualityStatus.REVIEW.value


def parse_csv(content: bytes) -> list[dict]:
    return list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))


def parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean(cell).lower() for cell in rows[0]]
    return [
        {headers[i]: value for i, value in enumerate(values) if i < len(headers)}
        for values in rows[1:]
    ]


def row_to_word(row: dict, default_active: bool) -> tuple[WordItem | None, str | None]:
    normalized = {k.strip().lower(): v for k, v in row.items() if k}
    word_text = _clean(normalized.get("word"))
    if not word_text:
        return None, "Missing word"

    values = {field: _get(normalized, field) for field in REQUIRED_FIELDS}
    missing = [field for field, value in values.items() if not value]
    if missing:
        return None, f"{word_text}: missing {', '.join(missing)}"

    quality_status = _quality_status(normalized, default_active)
    is_active = quality_status == WordQualityStatus.PUBLISHED.value

    return (
        WordItem(
            word=word_text,
            word_type=_get(normalized, "word_type") or "word",
            phonetic=_clean(normalized.get("phonetic")) or "-",
            level=_clean(normalized.get("level")) or "A1",
            topic=_clean(normalized.get("topic")) or "General",
            collection=_clean(normalized.get("collection")) or "Daily Vocabulary",
            tags=_clean(normalized.get("tags")),
            collocations=_clean(normalized.get("collocations")),
            common_mistake=_clean(normalized.get("common_mistake")),
            writing_prompt=_clean(normalized.get("writing_prompt")),
            difficulty_order=_to_int(normalized.get("difficulty_order")),
            audio_url=_clean(normalized.get("audio_url")) or None,
            audio_status="ready" if _clean(normalized.get("audio_url")) else "pending",
            quality_status=quality_status,
            is_active=is_active,
            **values,
        ),
        None,
    )


def import_rows(rows: list[dict], default_active: bool, db: Session) -> dict:
    imported = 0
    skipped: list[str] = []
    for line_number, row in enumerate(rows, start=2):
        word, error = row_to_word(row, default_active)
        if error or not word:
            skipped.append(f"Row {line_number}: {error or 'Invalid row'}")
            continue
        db.add(word)
        imported += 1
    db.commit()
    return {"imported": imported, "skipped": skipped[:25], "skipped_count": len(skipped)}
